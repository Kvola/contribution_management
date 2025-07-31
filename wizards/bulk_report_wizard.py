# -*- coding: utf-8 -*-

from odoo import models, fields, api
from odoo.exceptions import UserError, ValidationError
import logging
import base64

_logger = logging.getLogger(__name__)


class BulkReportWizard(models.TransientModel):
    """Wizard pour générer des rapports en lot"""
    
    _name = 'bulk.report.wizard'
    _description = 'Génération de rapports en lot'

    report_type = fields.Selection([
        ('member', 'Rapports membres'),
        ('group', 'Synthèses groupes'),
    ], string='Type de rapport', required=True, default='member')
    
    partner_ids = fields.Many2many(
        'res.partner',
        string='Partenaires',
        required=True,
        help="Sélectionnez les partenaires pour lesquels générer les rapports"
    )
    
    include_email = fields.Boolean(
        string='Envoyer par email',
        default=False,
        help="Envoyer automatiquement les rapports par email aux partenaires"
    )
    
    date_from = fields.Date(
        string='Date de début',
        help="Période de début pour le rapport (optionnel)"
    )
    
    date_to = fields.Date(
        string='Date de fin', 
        default=fields.Date.today(),
        help="Période de fin pour le rapport (optionnel)"
    )
    
    format_type = fields.Selection([
        ('pdf', 'PDF'),
        ('xlsx', 'Excel'),
    ], string='Format', default='pdf', required=True)
    
    separate_files = fields.Boolean(
        string='Fichiers séparés',
        default=True,
        help="Générer un fichier par partenaire, sinon un fichier consolidé"
    )

    @api.onchange('report_type')
    def _onchange_report_type(self):
        """Filtre les partenaires selon le type de rapport"""
        if self.report_type == 'group':
            domain = [('is_company', '=', True), ('active', '=', True)]
        else:
            domain = [('is_company', '=', False), ('active', '=', True)]
        
        return {'domain': {'partner_ids': domain}}

    @api.constrains('date_from', 'date_to')
    def _check_dates(self):
        """Valide les dates"""
        for wizard in self:
            if wizard.date_from and wizard.date_to and wizard.date_from > wizard.date_to:
                raise ValidationError("La date de début doit être antérieure à la date de fin.")

    def action_generate_reports(self):
        """Génère les rapports selon la configuration"""
        self.ensure_one()
        
        if not self.partner_ids:
            raise UserError("Veuillez sélectionner au moins un partenaire.")
        
        try:
            if self.format_type == 'pdf':
                return self._generate_pdf_reports()
            else:
                return self._generate_excel_reports()
                
        except Exception as e:
            _logger.error(f"Erreur lors de la génération des rapports en lot: {e}")
            raise UserError(f"Erreur lors de la génération des rapports: {e}")

    def _generate_pdf_reports(self):
        """Génère les rapports PDF"""
        if self.separate_files:
            return self._generate_separate_pdf_reports()
        else:
            return self._generate_consolidated_pdf_report()

    def _generate_separate_pdf_reports(self):
        """Génère des fichiers PDF séparés"""
        generated_reports = []
        email_sent_count = 0
        
        for partner in self.partner_ids:
            try:
                # Générer le rapport PDF
                if self.report_type == 'member':
                    report_ref = 'contribution_management.action_report_member_cotisations'
                else:
                    report_ref = 'contribution_management.action_report_group_synthesis'
                
                report_pdf = self.env.ref(report_ref).render_qweb_pdf([partner.id])
                
                if report_pdf and report_pdf[0]:
                    # Créer une pièce jointe
                    filename = self._get_report_filename(partner)
                    attachment = self.env['ir.attachment'].create({
                        'name': filename,
                        'type': 'binary',
                        'datas': base64.b64encode(report_pdf[0]),
                        'res_model': 'res.partner',
                        'res_id': partner.id,
                        'mimetype': 'application/pdf'
                    })
                    
                    generated_reports.append({
                        'partner': partner.name,
                        'attachment_id': attachment.id,
                        'filename': filename
                    })
                    
                    # Envoyer par email si demandé
                    if self.include_email and partner.email:
                        self._send_report_email(partner, attachment)
                        email_sent_count += 1
                        
            except Exception as e:
                _logger.error(f"Erreur lors de la génération du rapport pour {partner.name}: {e}")
                continue
        
        # Retourner le résumé
        return self._show_generation_summary(generated_reports, email_sent_count)

    def _generate_consolidated_pdf_report(self):
        """Génère un fichier PDF consolidé"""
        try:
            partner_ids = self.partner_ids.ids
            
            if self.report_type == 'member':
                report_ref = 'contribution_management.action_report_member_cotisations'
                filename = f'Rapports_membres_consolide_{fields.Date.today().strftime("%Y%m%d")}.pdf'
            else:
                report_ref = 'contribution_management.action_report_group_synthesis'
                filename = f'Syntheses_groupes_consolide_{fields.Date.today().strftime("%Y%m%d")}.pdf'
            
            # Générer le rapport PDF consolidé
            report_pdf = self.env.ref(report_ref).render_qweb_pdf(partner_ids)
            
            if report_pdf and report_pdf[0]:
                # Créer une pièce jointe
                attachment = self.env['ir.attachment'].create({
                    'name': filename,
                    'type': 'binary',
                    'datas': base64.b64encode(report_pdf[0]),
                    'res_model': 'bulk.report.wizard',
                    'res_id': self.id,
                    'mimetype': 'application/pdf'
                })
                
                # Retourner l'action de téléchargement
                return {
                    'type': 'ir.actions.act_url',
                    'url': f'/web/content/{attachment.id}?download=true',
                    'target': 'self',
                }
            
        except Exception as e:
            _logger.error(f"Erreur lors de la génération du rapport consolidé: {e}")
            raise UserError(f"Erreur lors de la génération du rapport consolidé: {e}")

    def _generate_excel_reports(self):
        """Génère les rapports Excel"""
        try:
            from openpyxl import Workbook
            from openpyxl.styles import Font, Alignment, PatternFill
            from io import BytesIO
            
            wb = Workbook()
            
            if self.separate_files:
                # Un fichier par partenaire
                return self._generate_separate_excel_reports(wb)
            else:
                # Un fichier consolidé
                return self._generate_consolidated_excel_report(wb)
                
        except ImportError:
            raise UserError("La bibliothèque openpyxl n'est pas disponible pour générer les fichiers Excel.")

    def _generate_separate_excel_reports(self, wb):
        """Génère des fichiers Excel séparés"""
        generated_reports = []
        
        for partner in self.partner_ids:
            try:
                # Créer un nouveau workbook pour chaque partenaire
                partner_wb = Workbook()
                
                if self.report_type == 'member':
                    self._create_member_excel_sheet(partner_wb, partner)
                else:
                    self._create_group_excel_sheet(partner_wb, partner)
                
                # Sauvegarder en mémoire
                buffer = BytesIO()
                partner_wb.save(buffer)
                buffer.seek(0)
                
                # Créer une pièce jointe
                filename = self._get_report_filename(partner, 'xlsx')
                attachment = self.env['ir.attachment'].create({
                    'name': filename,
                    'type': 'binary',
                    'datas': base64.b64encode(buffer.getvalue()),
                    'res_model': 'res.partner',
                    'res_id': partner.id,
                    'mimetype': 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
                })
                
                generated_reports.append({
                    'partner': partner.name,
                    'attachment_id': attachment.id,
                    'filename': filename
                })
                
            except Exception as e:
                _logger.error(f"Erreur lors de la génération Excel pour {partner.name}: {e}")
                continue
        
        return self._show_generation_summary(generated_reports, 0)

    def _create_member_excel_sheet(self, wb, member):
        """Crée une feuille Excel pour un membre"""
        ws = wb.active
        ws.title = f"Rapport {member.name[:20]}"
        
        # En-tête
        ws['A1'] = f"RAPPORT DE COTISATIONS - {member.name}"
        ws['A1'].font = Font(bold=True, size=14)
        ws['A1'].alignment = Alignment(horizontal='center')
        ws.merge_cells('A1:H1')
        
        # Informations membre
        row = 3
        ws[f'A{row}'] = "Membre:"
        ws[f'B{row}'] = member.name
        ws[f'D{row}'] = "Email:"
        ws[f'E{row}'] = member.email or 'Non défini'
        
        row += 1
        ws[f'A{row}'] = "Téléphone:"
        ws[f'B{row}'] = member.phone or 'Non défini'
        ws[f'D{row}'] = "Date rapport:"
        ws[f'E{row}'] = fields.Date.today().strftime('%d/%m/%Y')
        
        # Statistiques
        row += 3
        ws[f'A{row}'] = "STATISTIQUES"
        ws[f'A{row}'].font = Font(bold=True)
        
        row += 1
        stats = [
            ('Total cotisations:', member.total_cotisations),
            ('Payées:', member.paid_cotisations),
            ('En attente:', member.pending_cotisations),
            ('En retard:', member.overdue_cotisations),
            ('Taux paiement:', f"{member.payment_rate:.1f}%"),
        ]
        
        for label, value in stats:
            ws[f'A{row}'] = label
            ws[f'B{row}'] = value
            row += 1
        
        # Historique des cotisations
        row += 2
        ws[f'A{row}'] = "HISTORIQUE DES COTISATIONS"
        ws[f'A{row}'].font = Font(bold=True)
        
        row += 1
        headers = ['Type', 'Activité/Mois', 'Groupe', 'Montant dû', 'Montant payé', 'Date échéance', 'Date paiement', 'Statut']
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=row, column=col, value=header)
            cell.font = Font(bold=True)
            cell.fill = PatternFill(start_color="CCCCCC", end_color="CCCCCC", fill_type="solid")
        
        # Données des cotisations
        cotisations = member.cotisation_ids.filtered('active').sorted('due_date', reverse=True)
        for cotisation in cotisations:
            row += 1
            data = [
                'Activité' if cotisation.cotisation_type == 'activity' else 'Mensuelle',
                cotisation.activity_id.name if cotisation.activity_id else (
                    cotisation.monthly_cotisation_id.display_name if cotisation.monthly_cotisation_id else 'Non défini'
                ),
                cotisation.group_id.name if cotisation.group_id else 'Non défini',
                cotisation.amount_due,
                cotisation.amount_paid,
                cotisation.due_date.strftime('%d/%m/%Y') if cotisation.due_date else '-',
                cotisation.payment_date.strftime('%d/%m/%Y') if cotisation.payment_date else '-',
                self._get_state_label(cotisation.state)
            ]
            
            for col, value in enumerate(data, 1):
                ws.cell(row=row, column=col, value=value)
        
        # Ajuster la largeur des colonnes
        for column in ws.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)
            ws.column_dimensions[column_letter].width = adjusted_width

    def _create_group_excel_sheet(self, wb, group):
        """Crée une feuille Excel pour un groupe"""
        ws = wb.active
        ws.title = f"Synthèse {group.name[:20]}"
        
        # En-tête
        ws['A1'] = f"SYNTHÈSE DU GROUPE - {group.name}"
        ws['A1'].font = Font(bold=True, size=14)
        ws['A1'].alignment = Alignment(horizontal='center')
        ws.merge_cells('A1:H1')
        
        # Informations groupe
        row = 3
        ws[f'A{row}'] = "Groupe:"
        ws[f'B{row}'] = group.name
        ws[f'D{row}'] = "Email:"
        ws[f'E{row}'] = group.email or 'Non défini'
        
        row += 1
        ws[f'A{row}'] = "Téléphone:"
        ws[f'B{row}'] = group.phone or 'Non défini'
        ws[f'D{row}'] = "Date rapport:"
        ws[f'E{row}'] = fields.Date.today().strftime('%d/%m/%Y')
        
        # Statistiques globales
        row += 3
        ws[f'A{row}'] = "STATISTIQUES GLOBALES"
        ws[f'A{row}'].font = Font(bold=True)
        
        row += 1
        stats = [
            ('Membres total:', group.group_members_count),
            ('Membres actifs:', group.group_active_members_count),
            ('Activités:', group.activities_count),
            ('Cotisations mensuelles:', group.monthly_cotisations_count),
            ('Total collecté:', f"{group.group_total_collected:.2f}"),
            ('Total attendu:', f"{group.group_total_expected:.2f}"),
            ('Taux collecte:', f"{group.group_collection_rate:.1f}%"),
        ]
        
        for label, value in stats:
            ws[f'A{row}'] = label
            ws[f'B{row}'] = value
            row += 1
        
        # Liste des membres
        members = group.child_ids.filtered(lambda c: not c.is_company and c.active)
        if members:
            row += 2
            ws[f'A{row}'] = "MEMBRES DU GROUPE"
            ws[f'A{row}'].font = Font(bold=True)
            
            row += 1
            headers = ['Nom', 'Email', 'Téléphone', 'Cotisations', 'Payées', 'Taux paiement', 'Statut']
            for col, header in enumerate(headers, 1):
                cell = ws.cell(row=row, column=col, value=header)
                cell.font = Font(bold=True)
                cell.fill = PatternFill(start_color="CCCCCC", end_color="CCCCCC", fill_type="solid")
            
            # Données des membres
            for member in members.sorted('name'):
                row += 1
                status = 'Bon payeur' if member.is_good_payer else ('Retards' if member.has_overdue_payments else 'À surveiller')
                data = [
                    member.name,
                    member.email or '-',
                    member.phone or '-',
                    member.total_cotisations,
                    member.paid_cotisations,
                    f"{member.payment_rate:.1f}%",
                    status
                ]
                
                for col, value in enumerate(data, 1):
                    ws.cell(row=row, column=col, value=value)
        
        # Ajuster la largeur des colonnes
        for column in ws.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)
            ws.column_dimensions[column_letter].width = adjusted_width

    def _generate_consolidated_excel_report(self, wb):
        """Génère un fichier Excel consolidé"""
        try:
            # Supprimer la feuille par défaut
            wb.remove(wb.active)
            
            if self.report_type == 'member':
                # Créer une feuille pour chaque membre
                for partner in self.partner_ids:
                    if not partner.is_company:
                        ws = wb.create_sheet(title=partner.name[:31])  # Limite Excel: 31 caractères
                        self._create_member_excel_sheet_simple(ws, partner)
                        
                filename = f'Rapports_membres_consolide_{fields.Date.today().strftime("%Y%m%d")}.xlsx'
            else:
                # Créer une feuille pour chaque groupe
                for partner in self.partner_ids:
                    if partner.is_company:
                        ws = wb.create_sheet(title=partner.name[:31])
                        self._create_group_excel_sheet_simple(ws, partner)
                        
                filename = f'Syntheses_groupes_consolide_{fields.Date.today().strftime("%Y%m%d")}.xlsx'
            
            # Sauvegarder en mémoire
            buffer = BytesIO()
            wb.save(buffer)
            buffer.seek(0)
            
            # Créer une pièce jointe
            attachment = self.env['ir.attachment'].create({
                'name': filename,
                'type': 'binary',
                'datas': base64.b64encode(buffer.getvalue()),
                'res_model': 'bulk.report.wizard',
                'res_id': self.id,
                'mimetype': 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
            })
            
            # Retourner l'action de téléchargement
            return {
                'type': 'ir.actions.act_url',
                'url': f'/web/content/{attachment.id}?download=true',
                'target': 'self',
            }
            
        except Exception as e:
            _logger.error(f"Erreur lors de la génération Excel consolidée: {e}")
            raise UserError(f"Erreur lors de la génération Excel consolidée: {e}")

    def _create_member_excel_sheet_simple(self, ws, member):
        """Version simplifiée pour feuille consolidée"""
        # En-tête simplifié
        ws['A1'] = member.name
        ws['A1'].font = Font(bold=True, size=12)
        
        # Statistiques essentielles
        row = 3
        stats = [
            ('Total cotisations:', member.total_cotisations),
            ('Payées:', member.paid_cotisations),
            ('En retard:', member.overdue_cotisations),
            ('Taux paiement:', f"{member.payment_rate:.1f}%"),
        ]
        
        for label, value in stats:
            ws[f'A{row}'] = label
            ws[f'B{row}'] = value
            row += 1

    def _create_group_excel_sheet_simple(self, ws, group):
        """Version simplifiée pour feuille consolidée"""
        # En-tête simplifié
        ws['A1'] = group.name
        ws['A1'].font = Font(bold=True, size=12)
        
        # Statistiques essentielles
        row = 3
        stats = [
            ('Membres:', group.group_members_count),
            ('Activités:', group.activities_count),
            ('Total collecté:', f"{group.group_total_collected:.2f}"),
            ('Taux collecte:', f"{group.group_collection_rate:.1f}%"),
        ]
        
        for label, value in stats:
            ws[f'A{row}'] = label
            ws[f'B{row}'] = value
            row += 1

    def _get_report_filename(self, partner, extension='pdf'):
        """Génère le nom du fichier de rapport"""
        date_str = fields.Date.today().strftime('%Y%m%d')
        safe_name = partner.name.replace('/', '_').replace('\\', '_')[:50]
        
        if self.report_type == 'member':
            return f'Rapport_cotisations_{safe_name}_{date_str}.{extension}'
        else:
            return f'Synthese_groupe_{safe_name}_{date_str}.{extension}'

    def _get_state_label(self, state):
        """Retourne le libellé traduit de l'état"""
        labels = {
            'pending': 'En attente',
            'partial': 'Partiel',
            'paid': 'Payé',
            'overdue': 'En retard',
        }
        return labels.get(state, state)

    def _send_report_email(self, partner, attachment):
        """Envoie le rapport par email"""
        try:
            if self.report_type == 'member':
                template_ref = 'contribution_management.email_template_member_monthly_report'
                subject = f'Rapport de cotisations - {partner.name}'
            else:
                template_ref = 'contribution_management.email_template_group_monthly_report'
                subject = f'Synthèse du groupe - {partner.name}'
            
            mail_template = self.env.ref(template_ref, False)
            if mail_template:
                # Créer l'email avec la pièce jointe
                mail_values = mail_template.generate_email(partner.id)
                mail_values['attachment_ids'] = [(4, attachment.id)]
                
                # Envoyer l'email
                mail = self.env['mail.mail'].create(mail_values)
                mail.send()
                
                _logger.info(f"Rapport envoyé par email à {partner.name} ({partner.email})")
            
        except Exception as e:
            _logger.warning(f"Erreur lors de l'envoi email pour {partner.name}: {e}")

    def _show_generation_summary(self, generated_reports, email_sent_count):
        """Affiche le résumé de génération"""
        summary_lines = [
            f"Rapports générés: {len(generated_reports)}",
        ]
        
        if email_sent_count > 0:
            summary_lines.append(f"Emails envoyés: {email_sent_count}")
        
        # Créer la vue de résumé
        summary_text = "\n".join(summary_lines)
        
        # Si un seul rapport, le télécharger directement
        if len(generated_reports) == 1:
            attachment_id = generated_reports[0]['attachment_id']
            return {
                'type': 'ir.actions.act_url',
                'url': f'/web/content/{attachment_id}?download=true',
                'target': 'self',
            }
        
        # Sinon, afficher le résumé
        return {
            'type': 'ir.actions.client',
            'tag': 'display_notification',
            'params': {
                'title': 'Génération terminée',
                'message': summary_text,
                'type': 'success',
                'sticky': True,
            },
        }

    def action_preview_reports(self):
        """Prévisualise les rapports avant génération"""
        self.ensure_one()
        
        if not self.partner_ids:
            raise UserError("Veuillez sélectionner au moins un partenaire.")
        
        # Créer une vue de prévisualisation
        preview_data = []
        
        for partner in self.partner_ids:
            if self.report_type == 'member' and not partner.is_company:
                preview_data.append({
                    'name': partner.name,
                    'email': partner.email,
                    'cotisations': partner.total_cotisations,
                    'taux_paiement': partner.payment_rate,
                })
            elif self.report_type == 'group' and partner.is_company:
                preview_data.append({
                    'name': partner.name,
                    'email': partner.email,
                    'membres': partner.group_members_count,
                    'taux_collecte': partner.group_collection_rate,
                })
        
        return {
            'type': 'ir.actions.act_window',
            'name': 'Prévisualisation des rapports',
            'res_model': 'bulk.report.preview',
            'view_mode': 'tree',
            'target': 'new',
            'context': {
                'default_preview_data': preview_data,
                'default_report_type': self.report_type,
            },
        }


class BulkReportPreview(models.TransientModel):
    """Modèle pour prévisualiser les rapports en lot"""
    
    _name = 'bulk.report.preview'
    _description = 'Prévisualisation des rapports en lot'

    name = fields.Char(string='Nom', readonly=True)
    email = fields.Char(string='Email', readonly=True)
    cotisations = fields.Integer(string='Cotisations', readonly=True)
    taux_paiement = fields.Float(string='Taux paiement (%)', readonly=True)
    membres = fields.Integer(string='Membres', readonly=True)
    taux_collecte = fields.Float(string='Taux collecte (%)', readonly=True)
    report_type = fields.Char(string='Type', readonly=True)